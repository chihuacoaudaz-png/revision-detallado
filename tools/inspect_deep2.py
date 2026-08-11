"""
Script para:
1. Ver TODAS las filas del Maestro de Máquinas (Excepciones y NOMBRES_SAP)
2. Inspeccionar encabezados reales de los detallados (están en fila ~10-12 aprox.)
"""
import openpyxl
import os

# ============================================================
# 1. MAESTRO DE MÁQUINAS - COMPLETO
# ============================================================
print("=" * 80)
print("MAESTRO DE MÁQUINAS - TODAS LAS FILAS")
print("=" * 80)

maestro_path = r"c:\Proyectos Python\Detallados\Estructura base\Rockdrill_Control_Operaciones\Maestro_Maquinas\Maestros_Maquinas.xlsx"
wb2 = openpyxl.load_workbook(maestro_path, read_only=True, data_only=True)

for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    print(f"\n--- Hoja: '{sheet_name}' ---")
    for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
        row_data = list(row)
        if any(v is not None for v in row_data):
            # Limpiar None al final
            while row_data and row_data[-1] is None:
                row_data.pop()
            print(f"  Fila {i}: {row_data}")
wb2.close()

# ============================================================
# 2. ENCABEZADOS REALES DE UN DETALLADO (filas 1-12 de primera hoja)
# ============================================================
print("\n" + "=" * 80)
print("ESTRUCTURA INTERNA DE DETALLADOS (primeras 15 filas de primera hoja de máquina)")
print("=" * 80)

archivos_path = r"c:\Proyectos Python\Detallados\archivos"
# Inspeccionar solo 3 detallados representativos
samples = [
    "RD.402.P.01.F.01  Reporte Detallado de Avance COBRIZA - JULIO.xlsx",
    "RD.402.P.01.F.01 Reporte Detallado de Avance  CONDESTABLE -JULIO.xlsx",
    "RD.402.P.01.F.01 Reporte Detallado de Avance SAN CRISTOBAL -JULIO.xlsx",
    "RD 402 P 01 F 01 Reporte Detallado de Avance INMACULADA JULIO.xlsx",
    "RD.402.P.01.F.01 Reporte Detallado de Avance YAURICOCHA  - JULIO.xlsx",
]

for filename in samples:
    filepath = os.path.join(archivos_path, filename)
    if not os.path.exists(filepath):
        print(f"\n[NO ENCONTRADO] {filename}")
        continue
    
    print(f"\n{'='*60}")
    print(f"ARCHIVO: {filename}")
    print(f"{'='*60}")
    
    wb3 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    # Buscar hojas de máquina (no ADITIVOS, no GENERAL, no MÁQUINA N, no Tiempos, no Hoja1)
    machine_sheets = [s for s in wb3.sheetnames 
                      if s not in ['ADITIVOS', 'GENERAL', 'Tiempos', 'Hoja1', 'Hoja3', 'LISTAS']
                      and not s.startswith('MÁQUINA') and not s.startswith('Maquina')]
    
    print(f"  Todas las hojas: {wb3.sheetnames}")
    print(f"  Hojas de máquina: {machine_sheets}")
    
    # Inspeccionar la primera hoja de máquina
    if machine_sheets:
        first_sheet = machine_sheets[0]
        ws3 = wb3[first_sheet]
        print(f"\n  --- Inspección de hoja '{first_sheet}' (filas 1-15) ---")
        for i, row in enumerate(ws3.iter_rows(min_row=1, max_row=15, values_only=True), 1):
            row_data = list(row)
            # Filtrar valores no nulos con su posición
            non_null = [(j+1, v) for j, v in enumerate(row_data) if v is not None]
            if non_null:
                print(f"  Fila {i}: {non_null}")
    
    wb3.close()
