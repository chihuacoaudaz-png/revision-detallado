"""
Generador del Libro Excel Consolidador de Power Query (Entregable 2)
Rockdrill Group - Sistema de Consolidación Modular
"""
import os
import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Forzar UTF-8 en salida estándar de Windows
if sys.stdout.encoding != 'utf-8':
    try: sys.stdout.reconfigure(encoding='utf-8')
    except Exception: pass

def generar_excel_consolidador(
    ruta_salida: str = "output/CONSOLIDADOR_DETALLADOS_POWERQUERY.xlsx",
    ruta_fact_avance: str = "output/star_schema/fact_perforacion_avance.csv",
    ruta_anomalias: str = "output/reporte_anomalias_campo.xlsx"
) -> str:
    """
    Crea el libro Excel maestro configurado con parámetros visuales, hoja de consolidación
    y hoja de reporte de anomalías de campo para revisión inmediata de Jefatura y Administradoras.
    """
    out_path = Path(ruta_salida).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Quitar hoja por defecto
    
    # --------------------------------------------------------------------------
    # 1. HOJA DE PARÁMETROS DE POWER QUERY
    # --------------------------------------------------------------------------
    ws_params = wb.create_sheet(title="_PARAMETROS_")
    ws_params.views.sheetView[0].showGridLines = True
    
    # Estilos
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1F4E78")
    border_thin = Border(left=Side(style='thin', color='D9D9D9'),
                         right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'),
                         bottom=Side(style='thin', color='D9D9D9'))
    
    ws_params["B2"] = "⚙️ PARÁMETROS DE CONFIGURACIÓN POWER QUERY (MODIFICABLES)"
    ws_params["B2"].font = font_title
    
    param_headers = ["Nombre del Parámetro", "Tipo de Dato", "Valor Actual Configurado", "Descripción / Instrucción"]
    for col_idx, h in enumerate(param_headers, start=2):
        cell = ws_params.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    params_data = [
        ("RutaOrigenLocal", "Texto", r"C:\Proyectos Python\Detallados\data_in\CTRs", "Ruta de la carpeta local con las subcarpetas de los 19 CTRs"),
        ("TipoOrigen", "Texto (LOCAL/CLOUD)", "LOCAL", "Conmutador: Escribe 'LOCAL' para pruebas en disco o 'CLOUD' para SharePoint"),
        ("UrlSharePoint", "Texto (URL)", "https://rockdrillgroup.sharepoint.com/sites/Operaciones/Documentos Compartidos/CTRs", "URL de la biblioteca de SharePoint para producción en la nube"),
        ("ToleranciaHorasGuardia", "Decimal", 12.0, "Jornada estándar de guardia diaria en horas para balanceo"),
        ("MesOperativoCorte", "Entero", 25, "Día de corte minero oficial (Ciclo del 26 al 25)")
    ]
    
    for row_idx, p in enumerate(params_data, start=5):
        for col_idx, val in enumerate(p, start=2):
            cell = ws_params.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Segoe UI", size=10)
            cell.border = border_thin
            if col_idx == 4:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color="0070C0")
                cell.fill = PatternFill(start_color="F2F8FC", end_color="F2F8FC", fill_type="solid")

    # --------------------------------------------------------------------------
    # 2. HOJA DE CONSOLIDADO DE AVANCES (FACT AVANCE)
    # --------------------------------------------------------------------------
    ws_avances = wb.create_sheet(title="CONSOLIDADO_AVANCES")
    ws_avances.views.sheetView[0].showGridLines = True
    
    ws_avances["A1"] = "📊 CONSOLIDADO MODULAR DE AVANCES DE PERFORACIÓN (TODOS LOS CTRs)"
    ws_avances["A1"].font = font_title
    
    if os.path.exists(ruta_fact_avance):
        df_fa = pd.read_csv(ruta_fact_avance)
        # Escribir encabezados
        for col_idx, col_name in enumerate(df_fa.columns, start=1):
            cell = ws_avances.cell(row=3, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        # Escribir filas (muestra hasta 5000 filas para agilidad)
        for row_idx, r_vals in enumerate(df_fa.values[:5000], start=4):
            for col_idx, val in enumerate(r_vals, start=1):
                cell = ws_avances.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=9)
                cell.border = border_thin

    # --------------------------------------------------------------------------
    # 3. HOJA DE LOG DE ANOMALÍAS DE CAMPO
    # --------------------------------------------------------------------------
    ws_anom = wb.create_sheet(title="LOG_ANOMALIAS_CAMPO")
    ws_anom.views.sheetView[0].showGridLines = True
    
    ws_anom["A1"] = "🛡️ LOG DE ANOMALÍAS DE CAMPO DETECTADAS (SOLICITUD DE RECTIFICACIÓN A MINA)"
    ws_anom["A1"].font = Font(name="Segoe UI", size=14, bold=True, color="C00000")
    
    if os.path.exists(ruta_anomalias):
        df_an = pd.read_excel(ruta_anomalias)
        for col_idx, col_name in enumerate(df_an.columns, start=1):
            cell = ws_anom.cell(row=3, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            
        for row_idx, r_vals in enumerate(df_an.values, start=4):
            for col_idx, val in enumerate(r_vals, start=1):
                cell = ws_anom.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(name="Segoe UI", size=9)
                cell.border = border_thin

    # Autoajustar anchos de columnas
    for ws in [ws_params, ws_avances, ws_anom]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
            
    wb.save(str(out_path))
    print(f"✅ Libro Excel Consolidador creado exitosamente en: {out_path}")
    return str(out_path)

if __name__ == "__main__":
    generar_excel_consolidador()
