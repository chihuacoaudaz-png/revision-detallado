"""
Script para inspeccionar:
1. Encabezados de la base de comparación (julio2026.xlsx) - fila 1
2. Hojas y contenido del Maestro de Máquinas
3. Encabezados de cada detallado en la carpeta 'archivos'
"""
import openpyxl
import os
import json

# ============================================================
# 1. BASE DE COMPARACIÓN - julio2026.xlsx
# ============================================================
print("=" * 80)
print("1. BASE DE COMPARACIÓN - julio2026.xlsx")
print("=" * 80)

base_path = r"c:\Proyectos Python\Detallados\base de comparacion\julio2026.xlsx"
wb = openpyxl.load_workbook(base_path, read_only=True, data_only=True)
print(f"Hojas disponibles: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Hoja: '{sheet_name}' ---")
    # Leer fila 1 (encabezados)
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
    # Filtrar None
    headers_clean = [(i+1, h) for i, h in enumerate(headers) if h is not None]
    print(f"Total columnas con dato en fila 1: {len(headers_clean)}")
    for col_idx, header in headers_clean:
        print(f"  Col {col_idx}: {header}")

wb.close()

# ============================================================
# 2. MAESTRO DE MÁQUINAS
# ============================================================
print("\n" + "=" * 80)
print("2. MAESTRO DE MÁQUINAS - Maestros_Maquinas.xlsx")
print("=" * 80)

maestro_path = r"c:\Proyectos Python\Detallados\Estructura base\Rockdrill_Control_Operaciones\Maestro_Maquinas\Maestros_Maquinas.xlsx"
wb2 = openpyxl.load_workbook(maestro_path, read_only=True, data_only=True)
print(f"Hojas disponibles: {wb2.sheetnames}")

for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    print(f"\n--- Hoja: '{sheet_name}' ---")
    # Leer encabezados
    headers2 = []
    for row in ws2.iter_rows(min_row=1, max_row=1, values_only=True):
        headers2 = list(row)
    headers2_clean = [(i+1, h) for i, h in enumerate(headers2) if h is not None]
    print(f"Encabezados: {[h for _, h in headers2_clean]}")
    
    # Leer todas las filas para ver el contenido
    all_rows = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        row_data = list(row)
        if any(v is not None for v in row_data):
            all_rows.append(row_data)
    
    print(f"Total filas de datos: {len(all_rows)}")
    # Mostrar primeras 10 filas
    for i, row_data in enumerate(all_rows[:15]):
        print(f"  Fila {i+2}: {row_data}")
    if len(all_rows) > 15:
        print(f"  ... ({len(all_rows) - 15} filas más)")

wb2.close()

# ============================================================
# 3. ENCABEZADOS DE CADA DETALLADO
# ============================================================
print("\n" + "=" * 80)
print("3. ENCABEZADOS DE CADA DETALLADO (fila 1)")
print("=" * 80)

archivos_path = r"c:\Proyectos Python\Detallados\archivos"
for filename in sorted(os.listdir(archivos_path)):
    if filename.startswith("~$"):
        continue
    if not filename.lower().endswith(".xlsx"):
        continue
    
    filepath = os.path.join(archivos_path, filename)
    # Extraer nombre de contrato del nombre de archivo
    contract_name = filename.replace("RD.402.P.01.F.01", "").replace("RD 402 P 01 F 01", "").replace("G RD.402.P.01.F.01", "").replace("RD.402.P.01.F", "").replace("RRRD.402.P.01.F.01", "")
    contract_name = contract_name.replace("Reporte Detallado de Avance", "").replace("Reporte detallado de Avance", "").strip()
    contract_name = contract_name.replace(".xlsx", "").replace(".XLSX", "").strip(" -")
    
    print(f"\n--- {contract_name} ({filename}) ---")
    try:
        wb3 = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        print(f"  Hojas: {wb3.sheetnames}")
        
        # Solo leer la primera hoja
        ws3 = wb3[wb3.sheetnames[0]]
        headers3 = []
        for row in ws3.iter_rows(min_row=1, max_row=1, values_only=True):
            headers3 = list(row)
        headers3_clean = [(i+1, h) for i, h in enumerate(headers3) if h is not None]
        print(f"  Columnas con encabezado: {len(headers3_clean)}")
        for col_idx, header in headers3_clean:
            print(f"    Col {col_idx}: {header}")
        wb3.close()
    except Exception as e:
        print(f"  ERROR: {e}")
